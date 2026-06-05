from __future__ import annotations

import csv
import json
import shutil
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
PROCESSED = ROOT / "data" / "processed"
STORY_DATA = ROOT / "storytelling" / "assets" / "data" / "story_data.json"

PISA_CHAPTER_5_6_URL = "https://stat.link/wh9d4z"
TALIS_CHAPTER_1_URL = "https://stat.link/files/90df6235-en/fvn03o.xlsx"
TALIS_CHAPTER_4_URL = "https://stat.link/files/90df6235-en/0fcxr7.xlsx"
UNESCO_API = "https://data.unesco.org/api/explore/v2.1/catalog/datasets/uis001/records"

PISA_CHAPTER_5_6 = RAW / "pisa_2022_volume_1_chapters_5_6.xlsx"
TALIS_CHAPTER_1 = RAW / "talis_2024_chapter_1.xlsx"
TALIS_CHAPTER_4 = RAW / "talis_2024_chapter_4.xlsx"

YEARS = [2015, 2018, 2022]

COUNTRIES = {
    "ESP": {
        "name": "España",
        "pisa": "Spain",
        "talis": "Spain",
        "unesco": "ESP",
        "region": "Europa",
        "color": "#d1495b",
    },
    "PRT": {
        "name": "Portugal",
        "pisa": "Portugal",
        "talis": "Portugal",
        "unesco": "PRT",
        "region": "Europa",
        "color": "#edae49",
    },
    "FRA": {
        "name": "Francia",
        "pisa": "France",
        "talis": "France",
        "unesco": "FRA",
        "region": "Europa",
        "color": "#30638e",
    },
    "FIN": {
        "name": "Finlandia",
        "pisa": "Finland",
        "talis": "Finland",
        "unesco": "FIN",
        "region": "Europa",
        "color": "#00798c",
    },
    "EST": {
        "name": "Estonia",
        "pisa": "Estonia",
        "talis": "Estonia",
        "unesco": "EST",
        "region": "Europa",
        "color": "#2a9d8f",
    },
    "USA": {
        "name": "Estados Unidos",
        "pisa": "United States",
        "talis": "United States",
        "unesco": "USA",
        "region": "América",
        "color": "#7c3aed",
    },
    "BRA": {
        "name": "Brasil",
        "pisa": "Brazil",
        "talis": "Brazil",
        "unesco": "BRA",
        "region": "América",
        "color": "#15803d",
    },
    "CHL": {
        "name": "Chile",
        "pisa": "Chile",
        "talis": "Chile",
        "unesco": "CHL",
        "region": "América",
        "color": "#b45309",
    },
    "COL": {
        "name": "Colombia",
        "pisa": "Colombia",
        "talis": "Colombia",
        "unesco": "COL",
        "region": "América",
        "color": "#0f766e",
    },
    "JPN": {
        "name": "Japón",
        "pisa": "Japan",
        "talis": "Japan",
        "unesco": "JPN",
        "region": "Asia-Pacífico",
        "color": "#be123c",
    },
    "KOR": {
        "name": "Corea",
        "pisa": "Korea",
        "talis": "Korea",
        "unesco": "KOR",
        "region": "Asia-Pacífico",
        "color": "#2563eb",
    },
    "SGP": {
        "name": "Singapur",
        "pisa": "Singapore",
        "talis": "Singapore",
        "unesco": "SGP",
        "region": "Asia-Pacífico",
        "color": "#9333ea",
    },
    "AUS": {
        "name": "Australia",
        "pisa": "Australia",
        "talis": "Australia",
        "unesco": "AUS",
        "region": "Asia-Pacífico",
        "color": "#0284c7",
    },
    "ARE": {
        "name": "Emiratos Árabes Unidos",
        "pisa": "United Arab Emirates",
        "talis": "United Arab Emirates",
        "unesco": "ARE",
        "region": "Oriente Medio",
        "color": "#c2410c",
    },
}

SUBJECTS = {
    "matematicas": {
        "label": "Matemáticas",
        "sheet": "Table I.B1.5.4",
    },
    "lectura": {
        "label": "Lectura",
        "sheet": "Table I.B1.5.5",
    },
    "ciencias": {
        "label": "Ciencias",
        "sheet": "Table I.B1.5.6",
    },
}

AI_PRACTICES = {
    "resumir": {
        "label": "Resumir o aprender un tema",
        "var": "b.meanpct.d_tt4g37byes",
    },
    "planes": {
        "label": "Generar planes o actividades",
        "var": "b.meanpct.d_tt4g37cyes",
    },
    "adaptar": {
        "label": "Adaptar materiales por necesidad",
        "var": "b.meanpct.d_tt4g37eyes",
    },
    "necesidades": {
        "label": "Apoyar necesidades específicas",
        "var": "b.meanpct.d_tt4g37dyes",
    },
    "evaluar": {
        "label": "Evaluar o calificar trabajos",
        "var": "b.meanpct.d_tt4g37ayes",
    },
    "feedback": {
        "label": "Redactar feedback o comunicaciones",
        "var": "b.meanpct.d_tt4g37fyes",
    },
}

AI_BENEFITS = {
    "planes": {
        "label": "Mejora la planificación de clases",
        "var": "b.meanpct.d_tt4g35aagree",
    },
    "adaptacion": {
        "label": "Adapta materiales a capacidades distintas",
        "var": "b.meanpct.d_tt4g35bagree",
    },
    "individual": {
        "label": "Apoya al alumnado individualmente",
        "var": "b.meanpct.d_tt4g35cagree",
    },
    "necesidades": {
        "label": "Ayuda con necesidades específicas",
        "var": "b.meanpct.d_tt4g35dagree",
    },
    "administracion": {
        "label": "Automatiza tareas administrativas",
        "var": "b.meanpct.d_tt4g35eagree",
    },
}

AI_CHALLENGES = {
    "autoría": {
        "label": "Facilita presentar trabajo ajeno como propio",
        "var": "b.meanpct.d_tt4g35fagree",
    },
    "errores": {
        "label": "Puede recomendar contenido incorrecto",
        "var": "b.meanpct.d_tt4g35gagree",
    },
    "sesgos": {
        "label": "Amplifica sesgos o malentendidos",
        "var": "b.meanpct.d_tt4g35hagree",
    },
    "privacidad": {
        "label": "Compromete privacidad y seguridad",
        "var": "b.meanpct.d_tt4g35iagree",
    },
    "pedagogía": {
        "label": "Sugiere enfoques pedagógicos inadecuados",
        "var": "b.meanpct.d_tt4g35jagree",
    },
}

AI_BARRIERS = {
    "infraestructura": {
        "label": "Falta infraestructura en el centro",
        "var": "b.meanpct.d_tt4g38ayes",
    },
    "competencias": {
        "label": "Faltan conocimientos y habilidades",
        "var": "b.meanpct.d_tt4g38byes",
    },
    "rechazo": {
        "label": "No cree que deba usarse IA",
        "var": "b.meanpct.d_tt4g38cyes",
    },
    "normas": {
        "label": "El centro no permite usar IA",
        "var": "b.meanpct.d_tt4g38dyes",
    },
    "sobrecarga": {
        "label": "Siente sobrecarga ante nuevas tecnologías",
        "var": "b.meanpct.d_tt4g38eyes",
    },
}

DIGITAL_INDICATORS = {
    "internet_secundaria": {
        "label": "Centros de secundaria con internet pedagógico",
        "indicator_id": "SCHBSP.2T3.WINTERN",
    },
    "ordenadores_secundaria": {
        "label": "Centros de secundaria con ordenadores pedagógicos",
        "indicator_id": "SCHBSP.2T3.WCOMPUT",
    },
}


def ensure_dirs() -> None:
    RAW.mkdir(parents=True, exist_ok=True)
    PROCESSED.mkdir(parents=True, exist_ok=True)
    STORY_DATA.parent.mkdir(parents=True, exist_ok=True)


def download(url: str, destination: Path) -> None:
    if destination.exists():
        return
    request = urllib.request.Request(url, headers={"User-Agent": "PR2-visualizacion-datos/2.0"})
    with urllib.request.urlopen(request, timeout=180) as response:
        with destination.open("wb") as handle:
            shutil.copyfileobj(response, handle)


def clean_name(value: Any) -> str:
    return str(value).replace("*", "").strip() if value not in (None, "") else ""


def number(value: Any, digits: int = 1) -> float | None:
    if value in (None, "", "m", "a", "c", "#N/A"):
        return None
    try:
        return round(float(value), digits)
    except (TypeError, ValueError):
        return None


def country_code_from_field(field: str, value: str) -> str | None:
    cleaned = clean_name(value)
    for code, meta in COUNTRIES.items():
        if cleaned == meta[field]:
            return code
    return None


def find_year_columns(ws, wanted_years: list[int]) -> dict[int, int]:
    columns: dict[int, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(8, col).value
        subheader = ws.cell(9, col).value
        if not header or "Mean score" not in str(subheader):
            continue
        for year in wanted_years:
            if f"PISA {year}" in str(header):
                columns[year] = col
    return columns


def column_for_variable(ws, variable: str) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(11, col).value == variable:
            return col
    raise RuntimeError(f"Variable not found in {ws.title}: {variable}")


def iter_talis_country_rows(ws):
    for row_idx in range(12, ws.max_row + 1):
        name = clean_name(ws.cell(row_idx, 1).value)
        isced = ws.cell(row_idx, 2).value
        if not name or isced != 2:
            continue
        code = country_code_from_field("talis", name)
        if code:
            yield code, row_idx


def extract_pisa_scores() -> list[dict[str, Any]]:
    wb = load_workbook(PISA_CHAPTER_5_6, data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    for subject, config in SUBJECTS.items():
        ws = wb[config["sheet"]]
        columns = find_year_columns(ws, YEARS)
        for row_idx in range(10, ws.max_row + 1):
            code = country_code_from_field("pisa", ws.cell(row_idx, 1).value)
            if not code:
                continue
            for year, col in columns.items():
                rows.append(
                    {
                        "country_code": code,
                        "country": COUNTRIES[code]["name"],
                        "region": COUNTRIES[code]["region"],
                        "year": year,
                        "subject": subject,
                        "subject_label": config["label"],
                        "score": number(ws.cell(row_idx, col).value, 1),
                        "unit": "Puntuación media PISA",
                        "source": "OECD, PISA 2022 Database, Tables I.B1.5.4-I.B1.5.6",
                    }
                )
    return sorted(rows, key=lambda item: (item["country_code"], item["subject"], item["year"]))


def extract_talis_ai_use() -> tuple[list[dict[str, Any]], dict[str, float | None]]:
    wb = load_workbook(TALIS_CHAPTER_1, data_only=True, read_only=True)
    ws = wb["Table 1.59"]
    rows: list[dict[str, Any]] = []
    averages = {
        "talis_average": None,
        "oecd_average": None,
    }
    for row_idx in range(12, ws.max_row + 1):
        name = clean_name(ws.cell(row_idx, 1).value)
        value = number(ws.cell(row_idx, 3).value, 1)
        if name == "TALIS average-49":
            averages["talis_average"] = value
        if name == "OECD average-27":
            averages["oecd_average"] = value
    for code, row_idx in iter_talis_country_rows(ws):
        rows.append(
            {
                "country_code": code,
                "country": COUNTRIES[code]["name"],
                "region": COUNTRIES[code]["region"],
                "ai_use_pct": number(ws.cell(row_idx, 3).value, 1),
                "standard_error": number(ws.cell(row_idx, 4).value, 2),
                "level": "ISCED 2",
                "source": "OECD, TALIS 2024 Results, Table 1.59",
            }
        )
    return sorted(rows, key=lambda item: item["country"]), averages


def extract_talis_training() -> list[dict[str, Any]]:
    wb = load_workbook(TALIS_CHAPTER_4, data_only=True, read_only=True)
    ws_training = wb["Table 4.22"]
    ws_need = wb["Table 4.27"]
    training_col = column_for_variable(ws_training, "b.meanpct.d_tt4g21gyes")
    need_col = column_for_variable(ws_need, "b.meanpct.d_tt4g24ghigh")

    by_code: dict[str, dict[str, Any]] = {}
    for code, row_idx in iter_talis_country_rows(ws_training):
        by_code.setdefault(code, {})["training_ai_pct"] = number(
            ws_training.cell(row_idx, training_col).value, 1
        )
    for code, row_idx in iter_talis_country_rows(ws_need):
        by_code.setdefault(code, {})["high_need_ai_pct"] = number(
            ws_need.cell(row_idx, need_col).value, 1
        )

    rows: list[dict[str, Any]] = []
    for code, metrics in by_code.items():
        training = metrics.get("training_ai_pct")
        need = metrics.get("high_need_ai_pct")
        rows.append(
            {
                "country_code": code,
                "country": COUNTRIES[code]["name"],
                "region": COUNTRIES[code]["region"],
                "training_ai_pct": training,
                "high_need_ai_pct": need,
                "need_training_gap": None
                if training is None or need is None
                else round(need - training, 1),
                "level": "ISCED 2",
                "source": "OECD, TALIS 2024 Results, Tables 4.22 and 4.27",
            }
        )
    return sorted(rows, key=lambda item: item["country"])


def extract_talis_metric_set(sheet: str, variables: dict[str, dict[str, str]], metric_type: str):
    wb = load_workbook(TALIS_CHAPTER_1, data_only=True, read_only=True)
    ws = wb[sheet]
    columns = {key: column_for_variable(ws, spec["var"]) for key, spec in variables.items()}
    rows: list[dict[str, Any]] = []
    for code, row_idx in iter_talis_country_rows(ws):
        for key, col in columns.items():
            rows.append(
                {
                    "country_code": code,
                    "country": COUNTRIES[code]["name"],
                    "region": COUNTRIES[code]["region"],
                    "metric_type": metric_type,
                    "metric_key": key,
                    "metric_label": variables[key]["label"],
                    "value": number(ws.cell(row_idx, col).value, 1),
                    "level": "ISCED 2",
                    "source": f"OECD, TALIS 2024 Results, {sheet}",
                }
            )
    return rows


def latest_unesco_value(country_id: str, indicator_id: str) -> dict[str, Any]:
    params = {
        "where": f"indicator_id='{indicator_id}' and country_id='{country_id}'",
        "order_by": "year desc",
        "limit": "1",
    }
    url = f"{UNESCO_API}?{urllib.parse.urlencode(params)}"
    request = urllib.request.Request(url, headers={"User-Agent": "PR2-visualizacion-datos/2.0"})
    with urllib.request.urlopen(request, timeout=40) as response:
        payload = json.load(response)
    if not payload["results"]:
        return {"value": None, "year": None}
    record = payload["results"][0]
    return {
        "value": number(record.get("value"), 1),
        "year": int(record["year"]) if record.get("year") else None,
    }


def extract_unesco_digital() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for code, country in COUNTRIES.items():
        for key, indicator in DIGITAL_INDICATORS.items():
            latest = latest_unesco_value(country["unesco"], indicator["indicator_id"])
            rows.append(
                {
                    "country_code": code,
                    "country": country["name"],
                    "region": country["region"],
                    "indicator": key,
                    "indicator_label": indicator["label"],
                    "value": latest["value"],
                    "year": latest["year"],
                    "unit": "% de centros",
                    "source": "UNESCO UIS SDG 4 Education Indicators, indicator 4.a.1",
                }
            )
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise RuntimeError(f"No rows to write for {path}")
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def compact_country_metadata() -> list[dict[str, Any]]:
    return [
        {
            "code": code,
            "name": meta["name"],
            "region": meta["region"],
            "color": meta["color"],
        }
        for code, meta in COUNTRIES.items()
    ]


def write_quality_summary(
    pisa_scores: list[dict[str, Any]],
    ai_use: list[dict[str, Any]],
    ai_training: list[dict[str, Any]],
    digital: list[dict[str, Any]],
) -> None:
    missing_pisa = sum(1 for row in pisa_scores if row["score"] is None)
    missing_digital = sum(1 for row in digital if row["value"] is None)
    summary = f"""# Resumen de calidad de datos

## Cobertura

- Sistemas educativos incluidos: {len(COUNTRIES)}
- Registros en `pisa_scores.csv`: {len(pisa_scores)}
- Registros en `talis_ai_use.csv`: {len(ai_use)}
- Registros en `talis_ai_training.csv`: {len(ai_training)}
- Registros en `sdg_digital_infrastructure.csv`: {len(digital)}

## Valores no disponibles

- Puntuaciones PISA no disponibles: {missing_pisa}
- Indicadores UNESCO no disponibles: {missing_digital}

Los valores no disponibles se conservan como celdas vacías en CSV y como `null`
en JSON. No se interpolan ni se sustituyen por medias.

## Fuentes usadas

- OECD PISA 2022 Results, Volume I, StatLinks de los capítulos 5 y 6.
- OECD TALIS 2024 Results, tablas 1.59 a 1.63, 4.22 y 4.27.
- UNESCO UIS SDG 4 Education Indicators, indicador 4.a.1.

## Nota metodológica

La visualización compara adopción, percepciones, preparación docente,
infraestructura digital y resultados académicos agregados. No afirma causalidad
entre uso de IA y rendimiento PISA.
"""
    (PROCESSED / "resumen_calidad.md").write_text(summary, encoding="utf-8")


def main() -> None:
    ensure_dirs()
    download(PISA_CHAPTER_5_6_URL, PISA_CHAPTER_5_6)
    download(TALIS_CHAPTER_1_URL, TALIS_CHAPTER_1)
    download(TALIS_CHAPTER_4_URL, TALIS_CHAPTER_4)

    pisa_scores = extract_pisa_scores()
    ai_use, ai_averages = extract_talis_ai_use()
    ai_training = extract_talis_training()
    ai_practices = extract_talis_metric_set("Table 1.60", AI_PRACTICES, "practice")
    ai_benefits = extract_talis_metric_set("Table 1.61", AI_BENEFITS, "benefit")
    ai_challenges = extract_talis_metric_set("Table 1.62", AI_CHALLENGES, "challenge")
    ai_barriers = extract_talis_metric_set("Table 1.63", AI_BARRIERS, "barrier")
    digital = extract_unesco_digital()

    perceptions = ai_practices + ai_benefits + ai_challenges + ai_barriers

    write_csv(PROCESSED / "pisa_scores.csv", pisa_scores)
    write_csv(PROCESSED / "talis_ai_use.csv", ai_use)
    write_csv(PROCESSED / "talis_ai_training.csv", ai_training)
    write_csv(PROCESSED / "talis_ai_perceptions.csv", perceptions)
    write_csv(PROCESSED / "sdg_digital_infrastructure.csv", digital)
    write_quality_summary(pisa_scores, ai_use, ai_training, digital)

    payload = {
        "metadata": {
            "title": "IA y educación: adopción, preparación y resultados",
            "question": (
                "¿Cómo está entrando la inteligencia artificial en la educación y qué "
                "diferencias aparecen entre uso docente, formación, infraestructura "
                "digital y resultados académicos?"
            ),
            "generated_from": [
                PISA_CHAPTER_5_6_URL,
                TALIS_CHAPTER_1_URL,
                TALIS_CHAPTER_4_URL,
                UNESCO_API,
            ],
            "notes": [
                "No se interpolan datos faltantes.",
                "TALIS mide uso y percepciones docentes; PISA mide rendimiento estudiantil.",
                "La relación entre IA y resultados se presenta como exploratoria, no causal.",
            ],
            "ai_averages": ai_averages,
        },
        "countries": compact_country_metadata(),
        "subjects": [
            {"key": key, "label": config["label"], "unit": "Puntuación media PISA"}
            for key, config in SUBJECTS.items()
        ],
        "ai_use": ai_use,
        "ai_training": ai_training,
        "ai_perceptions": perceptions,
        "digital": digital,
        "pisa_scores": pisa_scores,
    }
    STORY_DATA.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
